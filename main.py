import openpyxl
import os
import random

path = input("Inserisci il percorso del database: ")

while "\"" in path:
    path = input("Inserisci il percorso del database senza le virgolette: ")

questionsAlreadyAsked = []
numberOfAnswers = 4

def getNumberOfQuestions():
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    i = 1
    while ws.cell(row=i,column=1).value!=None:
        i+=1
    return i-1

def takeQuestion():
    questionNumber = random.randint(1,getNumberOfQuestions())

    while questionNumber in questionsAlreadyAsked:
        questionNumber = random.randint(1,getNumberOfQuestions())
    
    if questionNumber not in questionsAlreadyAsked:
        questionsAlreadyAsked.append(questionNumber)

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    print(">> ",ws.cell(row=questionNumber,column=1).value)
    
    # Show answers 
    
    answersAlreadySelected = []
    answers = []

    i = 1
    while i<=numberOfAnswers:
        answerNumber = random.randint(1,numberOfAnswers)

        while answerNumber in answersAlreadySelected:
            answerNumber = random.randint(1,numberOfAnswers)
        
        answers.append(ws.cell(row=questionNumber,column=answerNumber+1).value)
        answersAlreadySelected.append(answerNumber)
        i+=1

    m = 1   
    for answer in answers:
        print("\n",m,"-",answer)
        m+=1

    correctAnswer = ws.cell(row=questionNumber,column=2).value
    correctAnswerNumber = answers.index(correctAnswer)+1
    
    return correctAnswerNumber
    

choice = "y"
questionProgressNumber = 0
numberOfCorrectAnswers = 0

while choice=="y" and questionProgressNumber<getNumberOfQuestions():
    # Inizitalization
    os.system("cls")
    questionProgressNumber+=1
    
    print("Simulazioni\n")

    print("Domanda",questionProgressNumber,"su",getNumberOfQuestions(),"totali\n")
    correctAnswerNumber = takeQuestion()

    answerNumber = "a"
    while answerNumber.isnumeric()==False:
        answerNumber = input("\nRisposta: ")

    if int(answerNumber)==correctAnswerNumber:
        print("Risposta corretta!")
        numberOfCorrectAnswers+=1
    else:
        print("Risposta errata! Risposta corretta: ",correctAnswerNumber)
    
    choice="a"
    while choice!="y" and choice!="n":
        choice = input("\nNuova domanda? [y/n]: ")

print("\nRisposte corrette: ",numberOfCorrectAnswers,"/",questionProgressNumber)

os.system("pause")

